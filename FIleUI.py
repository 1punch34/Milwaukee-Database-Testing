from tkinter import messagebox
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import scrolledtext
import os
import json
import FileUploader
import LogModule

class File:
    def __init__(self, file_path):
        self.file_name = os.path.basename(file_path)
        self.file_path = file_path
        self.is_favorite = False

class FileExplorerApp:
    def __init__(self, master):
        self.master = master
        master.title("Milwaukee Converter")

        # Get the screen width and height
        screen_width = master.winfo_screenwidth()
        screen_height = master.winfo_screenheight()

        # Set the window size and position
        window_width = 600
        window_height = 400
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        master.geometry(f"{window_width}x{window_height}+{x}+{y}")

        # Dictionary to store file paths and names
        self.file_dict = {}

        # Create frame for buttons and table
        self.buttonFrame = tk.Frame(master)
        self.buttonFrame.place(relx=0.05, rely=0.1, relwidth=0.2, relheight=1)

        # Button to open file explorer
        self.openFileExplorerButton = tk.Button(self.buttonFrame, text="Select File", command=self.open_file_explorer, width=20, height=2)
        self.openFileExplorerButton.pack(pady=10)

        # Button to open selected files
        self.convertSelectedFiles = tk.Button(self.buttonFrame, text="Convert File", command=self.convert_selected_file, width=20, height=2)
        self.convertSelectedFiles.pack(pady=12)

        self.convertSelectedFiles = tk.Button(self.buttonFrame, text="Upload File", command=self.upload_selected_file, width=20, height=2, background="firebrick2")
        self.convertSelectedFiles.pack(pady=12)

        self.dropdown_var = tk.StringVar()
        self.dropdown = ttk.Combobox(self.buttonFrame, textvariable=self.dropdown_var, values=["Milwaukee Product Information", "Milwaukee Price List"])
        self.dropdown.pack(pady = 12)
        # Placeholder button
        self.functionButton = tk.Button(self.buttonFrame, text="Open File", command=self.open_selected_file, width=20, height=2)
        self.functionButton.pack(pady=12)

        # Button to close the application
        self.closeButton = tk.Button(self.buttonFrame, text="Close", command=self.close_app, width=20, height=2)
        self.closeButton.pack(pady=80)

        # Create frame for table
        self.tableFrame = tk.Frame(master)
        self.tableFrame.place(relx=0.3, rely=0.1, relwidth=0.65, relheight=0.8)

        # Table to display opened files
        self.openedFilesWindow = ttk.Treeview(self.tableFrame, columns=("File Name",), show="headings", height=10)
        self.openedFilesWindow.heading("File Name", text="File Name")
        self.openedFilesWindow.pack(fill=tk.BOTH, expand=True)
        self.openedFilesWindow.bind("<Double-1>", self.on_double_click)
        self.openedFilesWindow.bind("<ButtonRelease-1>", self.select_file)
        self.openedFilesWindow.bind("<Button-3>", self.clear_selection)

        # Load favorite files from JSON file
        self.openFavoriteFiles()
    
    def getFileReader(self):
        return self.dropdown_var.get()
    
    def open_file_explorer(self):
        filenames = filedialog.askopenfilenames(initialdir="/", title="Select files")
        if filenames:
            print("Selected files:", filenames)
            for filename in filenames:
                self.update_recent_files(filename)

    def update_recent_files(self, file_path):
        # Create File object
        file = File(file_path)
        # Add file to dictionary
        self.file_dict[file.file_path] = file
        # Insert file name into treeview
        self.openedFilesWindow.insert("", "end", values=(file.file_name, file_path))

    def clear_selection(self, event):
        self.openedFilesWindow.selection_remove(self.openedFilesWindow.selection())

    def select_file(self, event):
        global selected_files
        items = self.openedFilesWindow.selection()
        selected_files = [self.openedFilesWindow.item(item, "values")[1] for item in items]
    
    def show_log_popup(self, log_data):
        popup = tk.Toplevel(self.master)
        popup.title("Log Data")
        text_area = scrolledtext.ScrolledText(popup, wrap=tk.WORD, width=40, height=10)
        text_area.pack(expand=True, fill="both")

        # Insert the log data into the text area
        text_area.insert(tk.END, log_data)

        # Optionally make the text area read-only
        text_area.config(state=tk.DISABLED)
    

    def upload_selected_file(self):
        if selected_files is not None:
            for file in selected_files:
                if file:
                    fileReader = self.getFileReader()
                    if fileReader:
                        print(f"Uploading file: {file}")
                        FileUploader.addFiletoDatabase(file, fileReader)
                        log_data = LogModule.LogData.getLogData()
                        self.show_log_popup(log_data)
                    else:
                        messagebox.showinfo("Error", "No File Reader Selected")

    def convert_selected_file(self):
        # try:
            if selected_files is not None:
                for selected_file in selected_files:
                    if selected_file:
                        print(f"Converting file: {selected_file}")
                        
                        output = self.file_transformer(selected_file)
                        self.update_recent_files(output)
                        messagebox.showinfo("Conversion Status", "Success")
                        
        # except NameError:
        #     messagebox.showerror("No files selected. Please select files before opening.")

    def on_double_click(self, event):
        items = self.openedFilesWindow.selection()

        for item in items:
            # Retrieve both file path and file object
            file_path = self.openedFilesWindow.item(item, "values")[1]
            file = self.file_dict.get(file_path)
            if file:
                file_name = file.file_name
                if file.is_favorite:
                    self.openedFilesWindow.set(item, "#1", file_name)
                    file.is_favorite = False
                    self.file_dict[file_path] = file
                else:
                    self.openedFilesWindow.set(item, "#1", "⭐ " + file_name)
                    file.is_favorite = True
                    self.file_dict[file_path] = file
        # Save favorite files to JSON file
        self.save_favorite_files()

    def open_selected_file(self):
        try:
            if selected_files is not None:
                for selected_file in selected_files:
                    if selected_file:
                        print("Opening selected file:", selected_file)
                        os.startfile(selected_file)
        except NameError:
            messagebox.showerror("No files selected. Please select files before opening.")
            
    def openFavoriteFiles(self):
        try:
            with open("favorite_files.json", "r") as f:
                data = json.load(f)
                favorite_files = data.get("files", [])
                for file_path in favorite_files:
                    file = File(file_path)
                    file.is_favorite = True
                    # Add file to dictionary
                    self.file_dict[file.file_path] = file
                    # Insert file name into treeview
                    self.openedFilesWindow.insert("", "end", values=("⭐ " + file.file_name, file_path))
        except Exception:
            pass
        

    def save_favorite_files(self):
        favorite_files = [file_path for file_path, file in self.file_dict.items() if file.is_favorite]
        data = {"files": favorite_files}
        with open("favorite_files.json", "w") as f:
            json.dump(data, f)
    

    def close_app(self):
        # Save favorite files before closing the application
        self.save_favorite_files()
        self.master.destroy()

    def file_transformer(self, file_path: str):

        output_path = file_path.replace(".xlsx", " Combined.xlsx")
        errorLog = []

        original = pd.ExcelFile(file_path)
        print('Excel object created')

        sheets = original.sheet_names
        if "FR Product Information" in sheets:
            sheets.remove("FR Product Information")
        if "Digital Assets FR" in sheets:
            sheets.remove("Digital Assets FR")

        productInfo_DF = pd.read_excel(original, 'Product Information', dtype=str, na_values=[''], keep_default_na=False)
        productInfo_DF.fillna('', inplace=True)
        print('productInfo_DF read')

        assets_DF = pd.read_excel(original, 'Digital Assets', dtype=str, na_values=[''], keep_default_na=False)
        assets_DF.fillna('', inplace=True)
        print('assets_DF read')

        info_assets_DF = pd.merge(productInfo_DF, assets_DF, how="left", on='MFG Part # (OEM)')
        info_assets_DF.set_index('MFG Part # (OEM)', inplace=True)
        print('info & assets merged')

        info_assets_DF.to_excel(output_path)
        print('info & assets written to excel')

        wb = openpyxl.load_workbook(output_path)
        outputSheet = wb['Sheet1']
        print('Workbook loaded')

        for sheet in sheets[2:]:
            #try:
            specific_DF = pd.read_excel(original, sheet, dtype=str, na_values=[''], 
                                        keep_default_na=False, index_col='MFG Part # (OEM)') #usecols=lambda col: col != 'UPC'
            specific_DF.fillna('', inplace=True)
            # print(sheet)

            

            specNames = specific_DF.columns.tolist()
            #print(specNames)

            for part_num, row in specific_DF.iterrows():

                specifications = []
                #print('\tWorking on ' + str(part_num))
                specValues = row.values.tolist()
                # print(specValues)

                try:
                    target_row_index = info_assets_DF.index.get_loc(str(part_num)) + 2
                    #print("Target row: " + str(target_row_index))
                    #print('\t\tMatch found at row ' + str(target_row_index))
                except Exception as e:
                    # print("Error getting target row index:")
                    # print(e)
                    errorLog.append(part_num)
                    target_row_index = outputSheet.max_row + 1
                    outputSheet.cell(row=target_row_index, column=1).value = part_num
                    outputSheet.cell(row=target_row_index, column=7).value = specValues[0]
                        
                if len(specNames) == len(specValues):
                    #print('\tspecNames and specValues have matching length of ' + str(len(specNames)))
                    for i in range(len(specNames)):
                        specifications.append(specNames[i])
                        specifications.append(specValues[i])
                    #print('\tAssembled spec list:')
                    #print(specifications)

                    for col, spec in enumerate(specifications, start=129):
                        if col > len(specifications) + 128:
                            break
                        outputSheet.cell(row=target_row_index, column=col).value = spec
                    #print('\t\tSpecs output to sheet')
            
            # except Exception as e:
            #     print('\n\n--------------ERROR ERROR ERROR---------------------\n\n')
            #     log_msg = 'Error with sheet: ' + sheet
            #     errorLog.append(log_msg)
            #     print(log_msg)
            #     print(f"An error occurred: {e}")
            #     continue
        
        wb.save(output_path)
        wb.close()
        print("Completed. Error Log:")
        print(errorLog)
        return output_path

if __name__ == "__main__":
    root = tk.Tk()
    app = FileExplorerApp(root)
    root.mainloop()