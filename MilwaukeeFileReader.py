import dateutil.parser
import openpyxl
import dateutil
import json
import openpyxl.workbook
import LogModule
from QueryModule import QueryModule
from InventoryRecord import InventoryRecord
from ProductDetailsRecord import ProductDetailsRecord
from UOMRecord import UOMRecord
import FileUploader

logger = LogModule.Logger(__name__).logger


def openFile(excelfile):
    workbook = openpyxl.open(excelfile)
    logger.info("FILE OPENED")
    return workbook

def get_headers(sheet):
    header_row = sheet[1]  # Assuming headers are in the first row
    headers = {}
    for idx, cell in enumerate(header_row):
        if cell.value:  # Only add non-empty headers
            headers[cell.value] = idx
    return headers

def getInventory(row, header_indices):
    try:
        MFG_Part_Number = row[header_indices['MFG Part # (OEM)']]
        GTINCode = row[header_indices['GTIN']]
        status_index = header_indices.get('Status', None)
        Status = row[status_index] if status_index is not None and row[status_index] else None
        Short_Description = row[header_indices['Short Description']]
        Brand = 'Milwaukee'
        CountryCode = row[header_indices['Country Code - 2 Character code']]
        subBrand = row[header_indices['Sub-Brand']]
        warranty = row[header_indices['Manufacturer Warranty']]
        recalled = row[header_indices['Has Item Been Recalled']] == 'Y'
        previoussku = row[header_indices['Previous MFG Model #']]
        previousUPC = row[header_indices['Previous UPC']]
        minOrderQty = row[header_indices['Minimum Order Quantity']]
        mulOrderQty = row[header_indices['Multiple Order Quantity']]
        orderUOM = row[header_indices['Order UOM']]

    except Exception as err:
        logger.error(f"HEADER ERROR in Inventory query :  {err}")
    try:
        Show_Online_Date = dateutil.parser.parse(row[header_indices['Show Online Date']]).strftime('%Y-%m-%d')
        Available_to_Ship_Date = dateutil.parser.parse(row[header_indices['Available to Ship Date']]).strftime('%Y-%m-%d')
    except Exception as err:
        Show_Online_Date = row[header_indices['Show Online Date']]
        Available_to_Ship_Date = row[header_indices['Available to Ship Date']]
        logger.error(f"Error parsing Dates{Show_Online_Date, Available_to_Ship_Date} for row with with MFG_Part_Num: {MFG_Part_Number}: {err}")

    inventoryRecord = InventoryRecord(
            MFG_Part_Number=MFG_Part_Number, GTIN=GTINCode, Status=Status, Short_Description=Short_Description, 
            Show_Online_Date=Show_Online_Date, Available_to_Ship_Date=Available_to_Ship_Date, Brand= Brand, CountryCode=CountryCode, SubBrand=subBrand, Warranty=warranty
            , Recalled=recalled, PreviousModelNo=previoussku, PreviousUPC=previousUPC, MinOrder=minOrderQty, MulOrder=mulOrderQty, OrderUOM=orderUOM
        )

    return inventoryRecord

def getproductInfo(row, header_indices):
    try:
        modelno = row[header_indices['MFG Part # (OEM)']]
        search_keywords = row[header_indices['Search Keywords']]
        product_name = row[header_indices['Product Name']]
        description = row[header_indices['Marketing Copy']]
        features_list = [
            row[header_indices[f'Feature - Benefit Bullet {i}']] for i in range(1, 22)  if header_indices.get(f'Feature - Benefit Bullet {i}') and row[header_indices[f'Feature - Benefit Bullet {i}']]
        ]
        features = json.dumps(features_list)
        
        # Create the includes JSON object, filtering out any null values or empty strings
        includes_list = [row[header_indices['Package Contents']] if header_indices.get('Package Contents') and row[header_indices['Package Contents']] else '']
        includes = json.dumps(includes_list)
    except Exception as err:
        logger.error("HEADER ERROR in productInfo query")
    
    productDetailsRecord = ProductDetailsRecord(modelno=modelno, search_keywords=search_keywords, product_name=product_name, 
                                               description=description, features=features, includes=includes)
    
    return productDetailsRecord

def getDigitalAssets(row, digitalassetsheader):
    try:
        modelno = row[digitalassetsheader['MFG Part # (OEM)']]
        video = row[digitalassetsheader['Product Review Video']] 
        safetyDataSheet = row[digitalassetsheader['Safety Data Sheet (SDS) - PDF']]
        
        # Find index for the first hero image
        imageStart = digitalassetsheader.get('Main Product Image')
        if imageStart is None:
            raise KeyError('Main Product Image header is missing')

        # Find the index for the last Detailed Product View image
        detailed_views_indices = [idx for header, idx in digitalassetsheader.items() if header.startswith('Detailed Product View')]
        if not detailed_views_indices:
            raise KeyError('No Detailed Product View headers found')
        imagesEnd = max(detailed_views_indices)
    except Exception as err:
        logger.error(f"HEADER ERROR in Digital Assets query : {err}")
    # Extract all images between the first hero image and the last detailed product view
    imagesList = row[imageStart:imagesEnd + 1]
    imagesList= [img for img in imagesList if img]

    # Construct the JSON object
    images = json.dumps(imagesList)

    productDetailsRecord = ProductDetailsRecord(modelno=modelno, video=video, sds=safetyDataSheet, images=images)
    return productDetailsRecord

def getUOMs(row, header_indices):
    mfg_part_no = row[header_indices['MFG Part # (OEM)']] 
    desc = row[header_indices['Short Description']]
    uoms = []
    Eachupc = row[header_indices['UPC']]
    package_height = row[header_indices['Package Height (In.)']]
    package_width = row[header_indices['Package Width (In.)']]
    package_depth = row[header_indices['Package Depth (In.)']]
    package_weight = row[header_indices['Package Weight (Lb.)']]
    package_quantity = row[header_indices['Net Package Quantity/Net Content']]
    uom = 'EA' if package_quantity == '1' else f'PK{package_quantity}'

    EachUOM = UOMRecord(mfg_part_no= mfg_part_no, desc= desc, uom=uom, upc= Eachupc, quantity= package_quantity, weight= package_weight, width= package_width, height= package_height, depth= package_depth)
    uoms.append(EachUOM)
    # Inner pack details
    inner_pack_upc = row[header_indices['Inner Pack GTIN']]
    inner_pack_quantity = row[header_indices['Inner Pack Quantity']]
    inner_pack_height = row[header_indices['Inner Pack Height (In.)']]
    inner_pack_width = row[header_indices['Inner Pack Width (In.)']]
    inner_pack_depth = row[header_indices['Inner Pack Depth (In.)']]
    inner_pack_weight = row[header_indices['Inner Pack Weight (Lb.)']]
    
    if inner_pack_quantity:
        IpUom = UOMRecord(mfg_part_no=mfg_part_no, desc=desc, quantity=inner_pack_quantity, upc= inner_pack_upc, uom= f'IP{inner_pack_quantity}', height = inner_pack_height, depth=inner_pack_depth, width=inner_pack_width, weight=inner_pack_weight)    
        uoms.append(IpUom)
    # Case details
    case_upc = row[header_indices['Case GTIN']]
    case_quantity = row[header_indices['Case Quantity']]
    case_height = row[header_indices['Case Height (In.)']]
    case_width = row[header_indices['Case Width (In.)']]
    case_depth = row[header_indices['Case Depth (In.)']]
    case_weight = row[header_indices['Case Weight (Lb.)']]
    
    if case_quantity:
        caseUOM = UOMRecord(mfg_part_no=mfg_part_no, desc=desc, quantity= case_quantity, uom=f'CASE{case_quantity}', upc = case_upc, weight= case_weight, height= case_height,
                            width= case_width, depth=case_depth)
        uoms.append(caseUOM)
    
    return uoms

def getSpecs(row, header_indices):
    MFG_Part_Number = row[header_indices['MFG Part # (OEM)']]
    specDict = {}
    for header_name, index in header_indices.items():
        # Add the header name and value to the dictionary
        if row[index] is None:
            continue
        specDict[header_name] = row[index]

    specDict.pop("GTIN", None)
    specs = json.dumps(specDict)

    productDetails = ProductDetailsRecord(modelno= MFG_Part_Number, specs=specs)
    return productDetails

def productInformationSheet(workbook, conn):
    logger.info("READING PRODUCT INFORMATION SHEET")
    productInformationSheet = workbook['Product Information']
    # Create a dictionary to map header names to column indices
    header_indices = get_headers(productInformationSheet)
    queryModule = QueryModule(conn)

    for row in productInformationSheet.iter_rows(min_row=2, max_row=productInformationSheet.max_row, values_only=True):
        MFG_Part_Number = row[header_indices['MFG Part # (OEM)']]
        if MFG_Part_Number is None:
            continue
        try:
            #Read Inventory values and run inventory query
            inventoryRecord = getInventory(row, header_indices)
            FileUploader.inventoryQuery(inventoryRecord, queryModule)
            #read product detail values and call productdetails query
            productDetailsRecord = getproductInfo(row, header_indices)
            FileUploader.productInfoQuery(productDetailsRecord, queryModule)
            #read uoms
            uomsRecords = getUOMs(row, header_indices)
            for uom in uomsRecords:
                FileUploader.UOMquery(uom, queryModule)
        except Exception as err:
            logger.error(f"Funciton Error @ProductInfoSheet: {err}")
        print("============================000000000000=============================  \n\n\n")

def digitalAssetsSheet(workbook, conn):
    logger.info("READING DIGITAL ASSETS SHEET")
    digitalAssetsSheet = workbook['Digital Assets']
    digitalAssetsHeaders_indices = get_headers(digitalAssetsSheet)
    queryModule = QueryModule(conn)
    for row in digitalAssetsSheet.iter_rows(min_row=2, max_row=digitalAssetsSheet.max_row, values_only= True):
        MFG_Part_Number = row[digitalAssetsHeaders_indices['MFG Part # (OEM)']]
        if MFG_Part_Number is None:
            continue            
        try:
            #Reads digital Assets and updates/inserts productdetailsrecord
            productDetailsRecord = getDigitalAssets(row, digitalAssetsHeaders_indices)
            FileUploader.productInfoQuery(productDetailsRecord, queryModule)
        except Exception as err:
            logger.error(f"Funciton Error @DigitalAssetsSheet: {err}")
        print("============================000000000000=============================  \n\n\n")

def specSheets(workbook, conn):

    logger.info("READING SPEC SHEETS")
    queryModule = QueryModule(conn)

    excluded_sheets = {"Product Information", "FR Product Information", "Digital Assets", "Digital Assets FR"}
    for sheet_name in workbook.sheetnames:
        if sheet_name in excluded_sheets:
            continue
        specSheet = workbook[sheet_name]
        # Create a dictionary to map header names to column indices
        specheader_indices = get_headers(specSheet)

        for row in specSheet.iter_rows(min_row=2, max_row=specSheet.max_row, values_only= True):
            MFG_Part_Number = row[specheader_indices['MFG Part # (OEM)']]
            if MFG_Part_Number is None:
                continue            
            try:
                productDetailsRecord = getSpecs(row, specheader_indices)
                FileUploader.productInfoQuery(productDetailsRecord, queryModule)
            except Exception as err:
                logger.error(f"Funciton Error @SpecSheet:{sheet_name} : {err}")
            print("============================000000000000=============================  \n\n\n")

def readFile(excelfile):

    workbook = openFile(excelfile)
    conn = FileUploader.connect()
    
    if conn is None or not conn.is_connected():
        logger.error("No Connection with Databse")
        return None
    
    productInformationSheet(workbook, conn)
    digitalAssetsSheet(workbook, conn)
    specSheets(workbook, conn)
    conn.close()